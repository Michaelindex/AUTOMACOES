import csv
from openpyxl import load_workbook

# Caminhos dos arquivos (ajuste se necessário)
csv_path = r"C:\Users\m.santos\OneDrive - Corporate\Área de Trabalho\oneitemimp_filtered.csv"
xlsx_path = r"C:\Users\m.santos\OneDrive - Corporate\Área de Trabalho\template.xlsx"

# Carrega o arquivo Excel (planilha final)
wb = load_workbook(xlsx_path)
ws = wb.active  # Se quiser usar uma aba específica: wb["NomeDaAba"]

# Abre o CSV (planilha inicial) e lê linha a linha
with open(csv_path, mode='r', newline='', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=',')
    
    # Pula a primeira linha do CSV (cabeçalho ou linha que não será usada)
    next(reader)
    
    # Vamos começar a escrever na planilha final a partir da linha 3
    row_in_template = 3
    
    # Para cada linha de dados no CSV, preenche as colunas na planilha final
    for row in reader:
        # Observação: o "row" aqui é uma lista com todas as colunas da linha do CSV,
        # onde row[0] = Coluna A, row[1] = Coluna B, ..., row[9] = Coluna J, etc.
        
        # CSV col J (índice 9)  => planilha final col B (coluna=2 em openpyxl)
        ws.cell(row=row_in_template, column=2).value = row[9]
        
        # CSV col P (índice 15) => planilha final col G (coluna=7)
        ws.cell(row=row_in_template, column=7).value = row[15]
        
        # CSV col AJ (índice 35) => planilha final col I (coluna=9)
        ws.cell(row=row_in_template, column=9).value = row[35]
        
        # CSV col AA (índice 26) => planilha final col Q (coluna=17)
        ws.cell(row=row_in_template, column=17).value = row[26]
        
        # CSV col AD (índice 29) => planilha final col X (coluna=24)
        ws.cell(row=row_in_template, column=24).value = row[29]
        
        # CSV col AF (índice 31) => planilha final col AE (coluna=31)
        ws.cell(row=row_in_template, column=31).value = row[31]
        
        # CSV col Y (índice 24)  => planilha final col AS (coluna=45)
        ws.cell(row=row_in_template, column=45).value = row[24]
        
        # Incrementa a linha de destino na planilha final
        row_in_template += 1

# Salva as alterações no template
wb.save(xlsx_path)
