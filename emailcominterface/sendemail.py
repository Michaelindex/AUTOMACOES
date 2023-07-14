import tkinter as tk
from tkinter import filedialog
import win32com.client as win32
import openpyxl

def selecionar_anexo():
    # Função para selecionar o arquivo de anexo
    filename = filedialog.askopenfilename()
    entry_anexo.delete(0, tk.END)  # Limpa o campo de entrada do anexo
    entry_anexo.insert(0, filename)  # Insere o nome do arquivo selecionado no campo de entrada

def enviar_email():
    # Função para enviar os e-mails
    outlook = win32.Dispatch("Outlook.Application")

    lista_email = entry_emails.get().split(',')  # Obtém a lista de e-mails do campo de entrada
    assunto = entry_assunto.get()  # Obtém o assunto do e-mail do campo de entrada
    corpo = entry_corpo.get("1.0", "end-1c")  # Obtém o corpo do e-mail do campo de entrada de texto
    anexo = entry_anexo.get()  # Obtém o caminho do anexo do campo de entrada

    for email in lista_email:
        message = outlook.CreateItem(0)
        message.To = email
        message.Subject = assunto
        message.Body = corpo

        if anexo:
            message.Attachments.Add(anexo)  # Adiciona o anexo ao e-mail, se houver

        message.Send()  # Envia o e-mail

    label_status["text"] = "E-mails enviados com sucesso!"  # Atualiza a etiqueta de status

    # Preencher a planilha do Excel
    workbook = openpyxl.load_workbook("notas-fiscais.xlsx")
    sheet = workbook.active

    # Preencher as informações na planilha
    informacoes = [
        [entry_rzsocial.get(), entry_data_enviada.get(), entry_numero_pc.get(), entry_numero_nf.get(), entry_valor.get(), entry_data_emissao.get(), entry_data_vencimento.get()],
        # Adicione mais informações conforme necessário
    ]

    last_row = sheet.max_row

    for row, info in enumerate(informacoes, start=1):
        for col, value in enumerate(info, start=1):
            sheet.cell(row=last_row + row, column=col).value = value

    # Salvar a planilha
    workbook.save("notas-fiscais.xlsx")
    workbook.close()

# Cria a janela principal
window = tk.Tk()
window.title("Enviar E-mails")
window.geometry("400x700")

# Cria os widgets
label_emails = tk.Label(window, text="E-mails:")  # Rótulo para os e-mails
label_emails.pack()
entry_emails = tk.Entry(window, width=40)  # Campo de entrada para os e-mails
entry_emails.pack()

label_assunto = tk.Label(window, text="Assunto:")  # Rótulo para o assunto
label_assunto.pack()
entry_assunto = tk.Entry(window, width=40)  # Campo de entrada para o assunto
entry_assunto.pack()

label_corpo = tk.Label(window, text="Corpo:")  # Rótulo para o corpo do e-mail
label_corpo.pack()
entry_corpo = tk.Text(window, width=40, height=6)  # Campo de entrada de texto para o corpo do e-mail
entry_corpo.pack()

label_anexo = tk.Label(window, text="Anexo:")  # Rótulo para o anexo
label_anexo.pack()
entry_anexo = tk.Entry(window, width=30)  # Campo de entrada para o caminho do anexo
entry_anexo.pack()

label_rzsocial = tk.Label(window, text="razão-social:")
label_rzsocial.pack()
entry_rzsocial = tk.Entry(window, width=40)
entry_rzsocial.pack()

label_data_enviada = tk.Label(window, text="data-enviada:")
label_data_enviada.pack()
entry_data_enviada = tk.Entry(window, width=40)
entry_data_enviada.pack()

label_numero_pc = tk.Label(window, text="numero-pc:")
label_numero_pc.pack()
entry_numero_pc = tk.Entry(window, width=40)
entry_numero_pc.pack()

label_numero_nf = tk.Label(window, text="numero-nf:")
label_numero_nf.pack()
entry_numero_nf = tk.Entry(window, width=40)
entry_numero_nf.pack()

label_valor = tk.Label(window, text="valor:")
label_valor.pack()
entry_valor = tk.Entry(window, width=40)
entry_valor.pack()

label_data_emissao = tk.Label(window, text="data-emissao:")
label_data_emissao.pack()
entry_data_emissao = tk.Entry(window, width=40)
entry_data_emissao.pack()

label_data_vencimento = tk.Label(window, text="data-vencimento:")
label_data_vencimento.pack()
entry_data_vencimento = tk.Entry(window, width=40)
entry_data_vencimento.pack()

button_selecionar_anexo = tk.Button(window, text="Selecionar Anexo", command=selecionar_anexo)  # Botão para selecionar o anexo
button_selecionar_anexo.pack()

button_enviar = tk.Button(window, text="Enviar", command=enviar_email)  # Botão para enviar os e-mails
button_enviar.pack()

label_status = tk.Label(window, text="")  # Etiqueta para exibir o status do envio dos e-mails
label_status.pack()

# Inicia o loop principal da janela
window.mainloop()
